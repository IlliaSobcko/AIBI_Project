"""
Standalone Excel Analyzer Module
Analyzes text reports and generates Excel analytics without external dependencies.
Can be run independently: python -m features.excel_analyzer
"""

import os
import re
from datetime import datetime
from pathlib import Path
from collections import Counter
from typing import List, Dict, Tuple, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportAnalyzer:
    """Analyzes text reports and extracts key business metrics."""

    def __init__(self, reports_folder: str = "reports"):
        """
        Initialize the analyzer.

        Args:
            reports_folder: Path to folder containing text reports
        """
        self.reports_folder = Path(reports_folder)
        self.data = []
        self.faqs_data = []

    def extract_client_name(self, text: str) -> Optional[str]:
        """Extract client name from report text."""
        # Try various patterns for client name extraction
        patterns = [
            r'(?:Client|Company|Account|Customer)[\s:]+([A-Za-z0-9\s\-&\.]+?)(?:\n|$)',
            r'^([A-Za-z0-9\s\-&\.]+?)(?:\s+(?:Report|Deal|Status))',
            r'(?:Name|Client Name)[\s:]+([A-Za-z0-9\s\-&\.]+?)(?:\n|$)',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()

        # Fallback: use first line if it looks like a name
        first_line = text.split('\n')[0].strip()
        if first_line and not any(keyword in first_line.lower() for keyword in ['report', 'analysis', 'summary']):
            return first_line

        return "Unknown"

    def extract_deal_status(self, text: str) -> str:
        """Extract deal status (Win/Loss) from report text."""
        text_lower = text.lower()

        # Strong indicators for Win
        win_indicators = [
            r'\bstatus[\s:]+win',
            r'\bdeal[\s:]+won',
            r'\b(?:success|won|closed|approved)',
            r'\b(?:deal status|result)[\s:]+(?:win|success|closed)',
        ]

        # Strong indicators for Loss
        loss_indicators = [
            r'\bstatus[\s:]+loss',
            r'\bdeal[\s:]+lost',
            r'\b(?:lost|failed|rejected|declined)',
            r'\b(?:deal status|result)[\s:]+(?:loss|failed)',
        ]

        for pattern in win_indicators:
            if re.search(pattern, text_lower):
                return "Win"

        for pattern in loss_indicators:
            if re.search(pattern, text_lower):
                return "Loss"

        return "Unknown"

    def extract_revenue(self, text: str) -> float:
        """Extract revenue amount from report text."""
        # Look for currency amounts
        patterns = [
            r'[\$][\s]*([0-9,]+(?:\.[0-9]{2})?)',
            r'(?:Revenue|Amount|Deal Value|Price)[\s:]+\$([0-9,]+(?:\.[0-9]{2})?)',
            r'(?:Revenue|Amount|Deal Value|Price)[\s:]+([0-9,]+(?:\.[0-9]{2})?)',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                amount_str = match.group(1).replace(',', '')
                try:
                    return float(amount_str)
                except ValueError:
                    pass

        return 0.0

    def extract_faqs(self, text: str) -> List[str]:
        """Extract FAQ references from report text."""
        # Look for FAQ patterns
        patterns = [
            r'(?:FAQ|Q\d+|Question)[\s:]*([^\n]+)',
            r'(?:Common Questions?|Frequently Asked)[\s:]*([^\n]+)',
            r'(?:Q&A|FAQ)[\s\-]+([^\n]+)',
        ]

        faqs = []
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            faqs.extend([match.strip() for match in matches if match.strip()])

        return faqs

    def load_and_analyze_reports(self) -> bool:
        """
        Load all text reports from the reports folder and analyze them.

        Returns:
            True if analysis was successful, False otherwise
        """
        if not self.reports_folder.exists():
            print(f"Warning: Reports folder '{self.reports_folder}' does not exist.")
            return False

        # Get all text files from reports folder
        report_files = list(self.reports_folder.glob('*.txt')) + \
                      list(self.reports_folder.glob('*.md')) + \
                      list(self.reports_folder.glob('*.text'))

        if not report_files:
            print(f"No report files found in '{self.reports_folder}'")
            return True  # Return True as this is not an error condition

        print(f"Found {len(report_files)} report file(s)")

        for report_file in report_files:
            try:
                with open(report_file, 'r', encoding='utf-8') as f:
                    text = f.read()

                # Extract data
                client_name = self.extract_client_name(text)
                deal_status = self.extract_deal_status(text)
                revenue = self.extract_revenue(text)
                faqs = self.extract_faqs(text)

                # Store data
                self.data.append({
                    'Client Name': client_name,
                    'Deal Status': deal_status,
                    'Revenue ($)': revenue,
                    'Report File': report_file.name
                })

                # Store FAQ data (only for wins)
                if deal_status == "Win" and faqs:
                    for faq in faqs:
                        self.faqs_data.append({
                            'FAQ': faq,
                            'Client': client_name,
                            'Revenue': revenue
                        })

                print(f"[OK] Analyzed: {report_file.name}")

            except Exception as e:
                print(f"[ERROR] Error reading {report_file.name}: {e}")

        return True

    def get_top_faqs(self, top_n: int = 5) -> List[Tuple[str, int]]:
        """
        Get top N FAQs that led to successful deals.

        Args:
            top_n: Number of top FAQs to return

        Returns:
            List of tuples (FAQ, count)
        """
        if not self.faqs_data:
            return []

        # Count FAQ occurrences
        faq_texts = [faq['FAQ'] for faq in self.faqs_data]
        faq_counter = Counter(faq_texts)

        return faq_counter.most_common(top_n)

    def get_summary_statistics(self) -> Dict:
        """Calculate summary statistics from analyzed data."""
        if not self.data:
            return {
                'total_deals': 0,
                'wins': 0,
                'losses': 0,
                'win_rate': 0.0,
                'total_revenue': 0.0,
                'average_deal_value': 0.0
            }

        total = len(self.data)
        wins = sum(1 for d in self.data if d['Deal Status'] == 'Win')
        losses = sum(1 for d in self.data if d['Deal Status'] == 'Loss')

        total_revenue = sum(d['Revenue ($)'] for d in self.data if d['Deal Status'] == 'Win')
        avg_deal_value = total_revenue / wins if wins > 0 else 0.0

        return {
            'total_deals': total,
            'wins': wins,
            'losses': losses,
            'win_rate': (wins / total * 100) if total > 0 else 0.0,
            'total_revenue': total_revenue,
            'average_deal_value': avg_deal_value
        }


def create_excel_report(analyzer: ReportAnalyzer, output_path: str = "analytics_report.xlsx") -> bool:
    """
    Create an Excel report from analyzed data.

    Args:
        analyzer: ReportAnalyzer instance with data
        output_path: Path where to save the Excel file

    Returns:
        True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE and not PANDAS_AVAILABLE:
        print("Error: Neither pandas nor openpyxl is installed.")
        print("Install with: pip install pandas openpyxl")
        return False

    # Create output directory if it doesn't exist
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if PANDAS_AVAILABLE and OPENPYXL_AVAILABLE:
        return _create_excel_with_pandas_openpyxl(analyzer, output_path)
    elif PANDAS_AVAILABLE:
        return _create_excel_with_pandas(analyzer, output_path)
    else:
        return _create_excel_with_openpyxl(analyzer, output_path)


def _create_excel_with_pandas_openpyxl(analyzer: ReportAnalyzer, output_path: str) -> bool:
    """Create Excel report using pandas and openpyxl for formatting."""
    try:
        # Create DataFrame from analyzed data
        df = pd.DataFrame(analyzer.data)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Deals', index=False)

            # Write summary statistics
            stats = analyzer.get_summary_statistics()
            summary_df = pd.DataFrame([stats])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Write top FAQs
            top_faqs = analyzer.get_top_faqs(5)
            if top_faqs:
                faq_df = pd.DataFrame(
                    [(faq, count) for faq, count in top_faqs],
                    columns=['FAQ', 'Occurrences in Wins']
                )
                faq_df.to_excel(writer, sheet_name='Top FAQs', index=False)

            # Format the workbook
            workbook = writer.book

            # Format Deals sheet
            deals_sheet = writer.sheets['Deals']
            _apply_formatting(deals_sheet, df.columns)

            # Format Summary sheet
            summary_sheet = writer.sheets['Summary']
            _apply_formatting(summary_sheet, summary_df.columns)

            # Format Top FAQs sheet if it exists
            if top_faqs and 'Top FAQs' in workbook.sheetnames:
                faqs_sheet = writer.sheets['Top FAQs']
                _apply_formatting(faqs_sheet, ['FAQ', 'Occurrences in Wins'])

        print(f"[OK] Excel report created: {output_path}")
        return True

    except Exception as e:
        print(f"[ERROR] Error creating Excel report: {e}")
        return False


def _create_excel_with_pandas(analyzer: ReportAnalyzer, output_path: str) -> bool:
    """Create Excel report using pandas only."""
    try:
        df = pd.DataFrame(analyzer.data)

        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Deals', index=False)

            stats = analyzer.get_summary_statistics()
            summary_df = pd.DataFrame([stats])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            top_faqs = analyzer.get_top_faqs(5)
            if top_faqs:
                faq_df = pd.DataFrame(
                    [(faq, count) for faq, count in top_faqs],
                    columns=['FAQ', 'Occurrences in Wins']
                )
                faq_df.to_excel(writer, sheet_name='Top FAQs', index=False)

        print(f"[OK] Excel report created: {output_path}")
        return True

    except Exception as e:
        print(f"[ERROR] Error creating Excel report: {e}")
        return False


def _create_excel_with_openpyxl(analyzer: ReportAnalyzer, output_path: str) -> bool:
    """Create Excel report using openpyxl only."""
    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Create Deals sheet
        deals_sheet = wb.create_sheet('Deals')
        if analyzer.data:
            headers = ['Client Name', 'Deal Status', 'Revenue ($)', 'Report File']
            deals_sheet.append(headers)

            for item in analyzer.data:
                deals_sheet.append([
                    item['Client Name'],
                    item['Deal Status'],
                    item['Revenue ($)'],
                    item['Report File']
                ])

            _apply_formatting(deals_sheet, headers)

        # Create Summary sheet
        summary_sheet = wb.create_sheet('Summary')
        stats = analyzer.get_summary_statistics()
        summary_headers = ['Total Deals', 'Wins', 'Losses', 'Win Rate (%)', 'Total Revenue ($)', 'Avg Deal Value ($)']
        summary_sheet.append(summary_headers)
        summary_sheet.append([
            stats['total_deals'],
            stats['wins'],
            stats['losses'],
            f"{stats['win_rate']:.2f}",
            f"{stats['total_revenue']:,.2f}",
            f"{stats['average_deal_value']:,.2f}"
        ])
        _apply_formatting(summary_sheet, summary_headers)

        # Create Top FAQs sheet
        top_faqs = analyzer.get_top_faqs(5)
        if top_faqs:
            faq_sheet = wb.create_sheet('Top FAQs')
            faq_headers = ['FAQ', 'Occurrences in Wins']
            faq_sheet.append(faq_headers)

            for faq, count in top_faqs:
                faq_sheet.append([faq, count])

            _apply_formatting(faq_sheet, faq_headers)

        wb.save(output_path)
        print(f"[OK] Excel report created: {output_path}")
        return True

    except Exception as e:
        print(f"[ERROR] Error creating Excel report: {e}")
        return False


def _apply_formatting(sheet, headers):
    """Apply basic formatting to a sheet."""
    try:
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        print(f"Warning: Could not apply formatting: {e}")


def run_analytics(reports_folder: str = "reports",
                 output_file: str = "analytics_report.xlsx") -> bool:
    """
    Main function to run the analytics pipeline.

    This is the primary entry point for the module.

    Args:
        reports_folder: Path to folder containing text reports (default: 'reports')
        output_file: Path where to save the Excel report (default: 'analytics_report.xlsx')

    Returns:
        True if analysis completed successfully, False otherwise

    Example:
        from features.excel_analyzer import run_analytics
        run_analytics(reports_folder='reports', output_file='my_analytics.xlsx')
    """
    print("\n" + "="*60)
    print("EXCEL ANALYZER - Starting Report Analysis")
    print("="*60 + "\n")

    # Initialize analyzer
    analyzer = ReportAnalyzer(reports_folder=reports_folder)

    # Load and analyze reports
    print(f"[LOAD] Loading reports from: {reports_folder}")
    if not analyzer.load_and_analyze_reports():
        print("Error: Failed to analyze reports")
        return False

    # Generate statistics
    print("\n[STATS] Generating Summary Statistics...")
    stats = analyzer.get_summary_statistics()

    print(f"\n  Total Deals: {stats['total_deals']}")
    print(f"  Wins: {stats['wins']}")
    print(f"  Losses: {stats['losses']}")
    print(f"  Win Rate: {stats['win_rate']:.2f}%")
    print(f"  Total Revenue: ${stats['total_revenue']:,.2f}")
    print(f"  Average Deal Value: ${stats['average_deal_value']:,.2f}")

    # Get top FAQs
    print("\n[TOP5] Top 5 FAQs Leading to Successful Deals:")
    top_faqs = analyzer.get_top_faqs(5)
    if top_faqs:
        for i, (faq, count) in enumerate(top_faqs, 1):
            print(f"  {i}. {faq} ({count} occurrence{'s' if count != 1 else ''})")
    else:
        print("  No FAQs found in winning deals")

    # Create Excel report
    print(f"\n[EXCEL] Creating Excel Report...")
    if create_excel_report(analyzer, output_file):
        print(f"\n[SUCCESS] Analysis Complete!")
        print(f"   Report saved to: {output_file}")
        return True
    else:
        print("\n[ERROR] Failed to create Excel report")
        return False


if __name__ == '__main__':
    # Run when executed directly
    run_analytics()
