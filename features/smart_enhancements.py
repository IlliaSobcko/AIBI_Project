"""
Smart Enhancements Module
- Knowledge Base: Weekly analysis of successful replies
- Finance Export: Generate Excel reports with stats and finances
- Auto-Booking: Automatic Google Calendar event creation
"""

import os
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
import json
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class KnowledgeBaseManager:
    """Analyzes successful replies and updates instructions"""

    def __init__(self, instructions_path: str = "instructions.txt"):
        self.instructions_path = Path(instructions_path)
        self.approved_replies_path = Path("approved_replies.json")
        self.last_analysis_path = Path("last_kb_analysis.txt")

    def log_approved_reply(self, chat_id: int, chat_title: str, original_message: str,
                          sent_reply: str, confidence: float):
        """Log an approved reply for future analysis"""
        # Load existing log
        if self.approved_replies_path.exists():
            with open(self.approved_replies_path, 'r', encoding='utf-8') as f:
                approved_log = json.load(f)
        else:
            approved_log = []

        # Add new entry
        approved_log.append({
            "timestamp": datetime.now().isoformat(),
            "chat_id": chat_id,
            "chat_title": chat_title,
            "original_message": original_message,
            "sent_reply": sent_reply,
            "confidence": confidence
        })

        # Save
        with open(self.approved_replies_path, 'w', encoding='utf-8') as f:
            json.dump(approved_log, f, ensure_ascii=False, indent=2)

        print(f"[KB] Logged approved reply for {chat_title}")

    async def weekly_analysis(self, ai_client):
        """Analyze approved replies and generate style/FAQ update"""
        # Check if week has passed since last analysis
        if self.last_analysis_path.exists():
            last_time = datetime.fromisoformat(self.last_analysis_path.read_text())
            if datetime.now() - last_time < timedelta(days=7):
                print(f"[KB] Last analysis was {(datetime.now() - last_time).days} days ago, skipping")
                return None

        # Load approved replies
        if not self.approved_replies_path.exists():
            print("[KB] No approved replies to analyze yet")
            return None

        with open(self.approved_replies_path, 'r', encoding='utf-8') as f:
            approved_log = json.load(f)

        if len(approved_log) < 5:
            print(f"[KB] Only {len(approved_log)} approved replies, need at least 5")
            return None

        # Prepare data for AI analysis
        recent_replies = approved_log[-50:]  # Last 50 approved replies
        analysis_prompt = f"""
Analyze these {len(recent_replies)} successful approved replies and extract:

1. COMMUNICATION STYLE PATTERNS:
   - Tone and formality level
   - Common phrases and expressions
   - Greeting/closing patterns

2. FREQUENT QUESTIONS & ANSWERS:
   - What questions appear repeatedly?
   - What are the successful responses?

3. KNOWLEDGE GAPS:
   - Topics where responses were less confident
   - Areas needing more detailed instructions

APPROVED REPLIES DATA:
{json.dumps(recent_replies, ensure_ascii=False, indent=2)}

Generate a concise update for instructions.txt that incorporates these learnings.
Format: Clear bullet points, ready to append to existing instructions.
"""

        try:
            # Generate update using AI
            update = await ai_client.generate_completion(analysis_prompt, model="gpt-4")

            # Append to instructions
            timestamp = datetime.now().strftime("%Y-%m-%d")
            with open(self.instructions_path, 'a', encoding='utf-8') as f:
                f.write(f"\n\n## AUTO-GENERATED UPDATE ({timestamp}) - Knowledge Base Analysis\n")
                f.write(update)
                f.write("\n")

            # Update last analysis time
            self.last_analysis_path.write_text(datetime.now().isoformat())

            print(f"[KB] [SUCCESS] Instructions updated with learnings from {len(recent_replies)} replies")
            return update

        except Exception as e:
            print(f"[KB] [ERROR] Analysis failed: {e}")
            return None


class FinanceExporter:
    """Generate Excel reports with chat stats and finances"""

    def __init__(self):
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)

    def extract_financial_data(self, messages: List[Dict]) -> List[Dict]:
        """Extract income/expense keywords from messages"""
        financial_keywords = {
            'income': ['оплачено', 'оплата', 'надійшло', 'перевів', 'paid', 'payment received'],
            'expense': ['ціна', 'вартість', 'купив', 'сплатив', 'price', 'cost', 'spent']
        }

        transactions = []

        for msg in messages:
            text = msg.get('text', '').lower()

            # Check for income keywords
            for keyword in financial_keywords['income']:
                if keyword in text:
                    # Try to extract amount
                    import re
                    amounts = re.findall(r'(\d+(?:\.\d+)?)\s*(?:грн|uah|\$|usd)', text)
                    if amounts:
                        transactions.append({
                            'date': msg.get('date'),
                            'chat': msg.get('chat_title'),
                            'type': 'Income',
                            'amount': float(amounts[0]),
                            'description': text[:100]
                        })
                    break

            # Check for expense keywords
            for keyword in financial_keywords['expense']:
                if keyword in text:
                    import re
                    amounts = re.findall(r'(\d+(?:\.\d+)?)\s*(?:грн|uah|\$|usd)', text)
                    if amounts:
                        transactions.append({
                            'date': msg.get('date'),
                            'chat': msg.get('chat_title'),
                            'type': 'Expense',
                            'amount': float(amounts[0]),
                            'description': text[:100]
                        })
                    break

        return transactions

    def generate_excel_report(self, chat_stats: List[Dict], messages: List[Dict]) -> str:
        """Generate comprehensive Excel report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.export_dir / f"aibi_report_{timestamp}.xlsx"

        wb = Workbook()

        # Sheet 1: Chat Statistics
        ws1 = wb.active
        ws1.title = "Chat Statistics"

        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Chat", "Total Messages", "Avg Response Time (min)", "Avg Confidence %", "Auto-Replies", "Manual Reviews"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row, stat in enumerate(chat_stats, 2):
            ws1.cell(row, 1, getattr(stat, 'chat_title', 'Unknown'))
            ws1.cell(row, 2, getattr(stat, 'total_messages', 0))
            ws1.cell(row, 3, getattr(stat, 'avg_response_time', 0))
            ws1.cell(row, 4, getattr(stat, 'avg_confidence', 0))
            ws1.cell(row, 5, getattr(stat, 'auto_replies', 0))
            ws1.cell(row, 6, getattr(stat, 'manual_reviews', 0))
        # Sheet 2: Financial Data
        ws2 = wb.create_sheet("Finances")

        transactions = self.extract_financial_data(messages)

        headers2 = ["Date", "Chat", "Type", "Amount", "Description"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        for row, trans in enumerate(transactions, 2):
            ws2.cell(row, 1, trans['date'])
            ws2.cell(row, 2, trans['chat'])
            ws2.cell(row, 3, trans['type'])
            ws2.cell(row, 4, trans['amount'])
            ws2.cell(row, 5, trans['description'])

        # Summary
        total_income = sum(t['amount'] for t in transactions if t['type'] == 'Income')
        total_expense = sum(t['amount'] for t in transactions if t['type'] == 'Expense')

        ws2.cell(len(transactions) + 3, 1, "TOTAL INCOME:")
        ws2.cell(len(transactions) + 3, 2, total_income)
        ws2.cell(len(transactions) + 4, 1, "TOTAL EXPENSE:")
        ws2.cell(len(transactions) + 4, 2, total_expense)
        ws2.cell(len(transactions) + 5, 1, "NET:")
        ws2.cell(len(transactions) + 5, 2, total_income - total_expense)

        wb.save(filename)
        print(f"[FINANCE] Report generated: {filename}")
        return str(filename)


class AutoBookingManager:
    """Automatic Google Calendar event creation for meetings"""

    def __init__(self, calendar_client):
        self.calendar = calendar_client

    async def detect_meeting_request(self, message_text: str, ai_response: str) -> Optional[Dict]:
        """Detect if message contains meeting request"""
        meeting_keywords = [
            'зустріч', 'встреча', 'meeting', 'зустрінемося', 'побачимось',
            'приїхати', 'прийти', 'коли можна', 'коли вільний'
        ]

        text_lower = message_text.lower()
        has_meeting_keyword = any(kw in text_lower for kw in meeting_keywords)

        if not has_meeting_keyword:
            return None

        # Try to extract date/time
        import re
        from dateutil import parser

        # Common patterns
        date_patterns = [
            r'(\d{1,2}[./]\d{1,2})',  # DD.MM or DD/MM
            r'(завтра|сьогодні|післязавтра)',  # Ukrainian
            r'(tomorrow|today)',  # English
            r'(\d{1,2}\s+(?:січня|лютого|березня|квітня|травня|червня|липня|серпня|вересня|жовтня|листопада|грудня))'
        ]

        time_patterns = [
            r'(\d{1,2}:\d{2})',  # HH:MM
            r'(\d{1,2}\s*годин)',  # X година
            r'(ранку|вечора|дня)'  # Morning/evening
        ]

        extracted_info = {
            'has_meeting': True,
            'date_mentioned': False,
            'time_mentioned': False,
            'raw_text': message_text
        }

        for pattern in date_patterns:
            if re.search(pattern, text_lower):
                extracted_info['date_mentioned'] = True
                break

        for pattern in time_patterns:
            if re.search(pattern, text_lower):
                extracted_info['time_mentioned'] = True
                break

        return extracted_info

    async def create_pending_meeting(self, chat_id: int, chat_title: str, message_text: str,
                                    meeting_info: Dict) -> str:
        """Create 'Pending Meeting' event in Google Calendar"""
        # Default to tomorrow 10:00 if no time specified
        if meeting_info.get('date_mentioned'):
            # Try to parse from text
            start_time = datetime.now() + timedelta(days=1)
            start_time = start_time.replace(hour=10, minute=0, second=0)
        else:
            start_time = datetime.now() + timedelta(days=1)
            start_time = start_time.replace(hour=10, minute=0, second=0)

        event_title = f"[PENDING] Meeting with {chat_title}"
        event_description = f"""
AUTO-CREATED MEETING REQUEST

Client: {chat_title}
Chat ID: {chat_id}
Original Message: {message_text[:200]}

Status: PENDING CONFIRMATION
Please confirm time with client!
"""

        try:
            # GoogleCalendarClient.create_event signature:
            # create_event(summary, description, start_time, duration_minutes, calendar_id)
            event_id = self.calendar.create_event(
                summary=event_title,
                description=event_description,
                start_time=start_time,
                duration_minutes=60,
                calendar_id='primary'
            )

            print(f"[AUTO-BOOK] Created pending meeting for {chat_title}: {event_id}")
            return event_id
        except Exception as e:
            print(f"[AUTO-BOOK] [ERROR] Failed to create event: {e}")
            return None


# Initialize managers
knowledge_base = KnowledgeBaseManager()
finance_exporter = FinanceExporter()
# auto_booking will be initialized with calendar_client when available
