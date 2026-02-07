"""
Excel Reporting Module - Production-ready data collection and export.
Collects analytics from reports folder and prepares data for Excel export.
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from collections import defaultdict


# ============================================================================
# EXCEL DATA COLLECTOR
# ============================================================================

class ExcelDataCollector:
    """
    Collects and aggregates data from report files for Excel export.
    Handles:
    - Customer questions extraction
    - Confidence score analytics
    - Revenue and expense tracking
    - Chat statistics
    """

    def __init__(self, reports_dir: str = "reports"):
        """Initialize collector with reports directory"""
        self.reports_dir = Path(reports_dir)
        self.data = {
            "total_chats": 0,
            "chats": [],
            "customer_questions": [],
            "confidence_scores": [],
            "revenue_entries": [],
            "business_expenses": [],
            "auto_replies": 0,
            "drafts": 0,
            "high_confidence": 0
        }

    # ========================================================================
    # DATA COLLECTION
    # ========================================================================

    def collect_all_data(self) -> Dict:
        """
        Collect all data from reports folder.

        Returns:
            Dictionary with all collected analytics
        """
        if not self.reports_dir.exists():
            print(f"[EXCEL] Reports directory not found: {self.reports_dir}")
            return self.data

        report_files = list(self.reports_dir.glob("*.txt"))
        print(f"[EXCEL] Found {len(report_files)} report files")

        for report_file in report_files:
            try:
                self._process_report_file(report_file)
            except Exception as e:
                print(f"[ERROR] Failed to process {report_file.name}: {e}")

        return self.data

    def _process_report_file(self, file_path: Path):
        """Process a single report file"""
        content = file_path.read_text(encoding="utf-8", errors="ignore")
        lines = content.split("\n")

        chat_title = file_path.stem  # Filename without .txt
        self.data["total_chats"] += 1
        self.data["chats"].append({
            "title": chat_title,
            "file": file_path.name,
            "processed_at": datetime.now().isoformat()
        })

        # Parse report content
        for i, line in enumerate(lines):
            # Extract confidence score
            if "ВПЕВНЕНІСТЬ ШІ:" in line:
                try:
                    confidence = int(line.split(":")[1].strip())
                    self.data["confidence_scores"].append(confidence)

                    if confidence >= 80:
                        self.data["high_confidence"] += 1

                except (ValueError, IndexError):
                    pass

            # Count auto-replies
            elif "AUTO-REPLY SENT" in line:
                self.data["auto_replies"] += 1

            # Count drafts
            elif "DRAFT FOR REVIEW" in line:
                self.data["drafts"] += 1

            # Collect questions/insights
            elif any(keyword in line.lower() for keyword in ["питання", "запитання", "вопрос"]):
                if line.strip() and not line.startswith("["):
                    self.data["customer_questions"].append(line.strip())

    # ========================================================================
    # STATISTICS CALCULATION
    # ========================================================================

    def calculate_statistics(self) -> Dict:
        """Calculate analytics statistics"""
        confidence_scores = self.data["confidence_scores"]

        avg_confidence = (
            sum(confidence_scores) / len(confidence_scores)
            if confidence_scores
            else 0
        )

        return {
            "total_chats": self.data["total_chats"],
            "avg_confidence": round(avg_confidence, 1),
            "high_confidence_count": self.data["high_confidence"],
            "auto_replies": self.data["auto_replies"],
            "drafts": self.data["drafts"],
            "total_responses": self.data["auto_replies"] + self.data["drafts"],
            "unique_questions": len(set(self.data["customer_questions"])),
            "min_confidence": min(confidence_scores) if confidence_scores else 0,
            "max_confidence": max(confidence_scores) if confidence_scores else 0
        }

    # ========================================================================
    # DATA FORMATTING
    # ========================================================================

    def format_for_summary(self) -> str:
        """Format collected data for user display"""
        stats = self.calculate_statistics()

        summary = f"""
📊 **EXCEL EXPORT DATA SUMMARY**

📈 Chat Statistics:
  • Total Chats: {stats['total_chats']}
  • Average Confidence: {stats['avg_confidence']}%
  • High Confidence (≥80%): {stats['high_confidence_count']}
  • Min/Max Confidence: {stats['min_confidence']}% / {stats['max_confidence']}%

📤 Response Statistics:
  • Auto-Replies Sent: {stats['auto_replies']}
  • Drafts for Review: {stats['drafts']}
  • Total Responses: {stats['total_responses']}

💬 Question Analytics:
  • Customer Questions: {len(self.data['customer_questions'])}
  • Unique Questions: {stats['unique_questions']}

✅ Excel data collection complete and ready for export
"""
        return summary.strip()

    # ========================================================================
    # EXCEL EXPORT PREPARATION (for future openpyxl integration)
    # ========================================================================

    def prepare_excel_sheets(self) -> Dict:
        """
        Prepare data structure for Excel sheets.

        Returns:
            Dictionary with sheet data ready for openpyxl
        """
        stats = self.calculate_statistics()

        sheets = {
            "Summary": self._prepare_summary_sheet(stats),
            "Chat Analytics": self._prepare_chat_analytics(stats),
            "Confidence Scores": self._prepare_confidence_sheet(),
            "Questions": self._prepare_questions_sheet()
        }

        return sheets

    def _prepare_summary_sheet(self, stats: Dict) -> List[List]:
        """Prepare summary sheet data"""
        return [
            ["AIBI Analytics Report", "", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""],
            ["", "", ""],
            ["Metric", "Value", "Unit"],
            ["Total Chats Processed", stats['total_chats'], "chats"],
            ["Average Confidence", stats['avg_confidence'], "%"],
            ["High Confidence Count", stats['high_confidence_count'], "chats"],
            ["Auto-Replies Sent", stats['auto_replies'], "replies"],
            ["Drafts for Review", stats['drafts'], "drafts"],
            ["Unique Questions", stats['unique_questions'], "questions"],
            ["Min Confidence", stats['min_confidence'], "%"],
            ["Max Confidence", stats['max_confidence'], "%"],
        ]

    def _prepare_chat_analytics(self, stats: Dict) -> List[List]:
        """Prepare chat analytics sheet"""
        rows = [
            ["Chat Title", "File Name", "Processed At"],
        ]

        for chat in self.data["chats"]:
            rows.append([
                chat["title"],
                chat["file"],
                chat["processed_at"]
            ])

        return rows

    def _prepare_confidence_sheet(self) -> List[List]:
        """Prepare confidence scores sheet"""
        rows = [["Chat Number", "Confidence Score (%)"]]

        for i, score in enumerate(self.data["confidence_scores"], 1):
            rows.append([i, score])

        return rows

    def _prepare_questions_sheet(self) -> List[List]:
        """Prepare questions sheet"""
        rows = [["Question / Insight"]]

        unique_questions = set(self.data["customer_questions"])
        for question in sorted(unique_questions):
            rows.append([question])

        return rows

    # ========================================================================
    # EXPORT METHODS (Ready for future enhancement)
    # ========================================================================

    def export_to_excel(self, filename: str = "AIBI_Report.xlsx") -> Optional[str]:
        """
        Export data to Excel file (requires openpyxl).

        Args:
            filename: Output filename

        Returns:
            Path to created file, or None if failed
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            sheets_data = self.prepare_excel_sheets()

            for sheet_name, data in sheets_data.items():
                ws = wb.create_sheet(sheet_name)

                # Add data to worksheet
                for row_idx, row_data in enumerate(data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)

                        # Format header row
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Save file
            output_path = Path(filename)
            wb.save(output_path)

            print(f"[EXCEL] ✓ File created: {output_path.absolute()}")
            return str(output_path.absolute())

        except ImportError:
            print("[EXCEL] ⚠ openpyxl not installed. Install with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"[ERROR] Excel export failed: {e}")
            return None

    # ========================================================================
    # UTILITY METHODS
    # ========================================================================

    def get_summary(self) -> str:
        """Get formatted data summary"""
        self.collect_all_data()
        return self.format_for_summary()

    def reset(self):
        """Reset collected data"""
        self.data = {
            "total_chats": 0,
            "chats": [],
            "customer_questions": [],
            "confidence_scores": [],
            "revenue_entries": [],
            "business_expenses": [],
            "auto_replies": 0,
            "drafts": 0,
            "high_confidence": 0
        }


# ============================================================================
# SINGLETON INSTANCE
# ============================================================================

_collector_instance: Optional[ExcelDataCollector] = None


def get_excel_collector() -> ExcelDataCollector:
    """Get or create Excel data collector singleton"""
    global _collector_instance

    if _collector_instance is None:
        _collector_instance = ExcelDataCollector()

    return _collector_instance
